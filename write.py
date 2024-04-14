
from openpyxl import load_workbook, Workbook
from typing import List, Tuple


def write_to_excel(data: List[Tuple[str, float]], x_values: List[float], y_values: List[float], filename: str):

    """
    Write data, x_values, and y_values to an Excel file.

    Args:
        data (List[Tuple[str, float]]): List of tuples containing attribute names and corresponding values.
        x_values (List[float]): List of x values.
        y_values (List[float]): List of y values.
        filename (str): Name of the Excel file to write to.
    """
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    # Find the next empty row after leaving three empty rows
    empty_rows_count = 0
    next_row = 1  # Start from the first row
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            empty_rows_count += 1
            if empty_rows_count == 3:
                next_row = ws.max_row + 1  # Next empty row after leaving three empty rows
                break
        else:
            empty_rows_count = 0  # Reset the count if a non-empty row is encountered

    # Write data to the worksheet starting from the next empty row
    for row_data in data:
        ws.append(row_data)

    # Write x_values and y_values to separate cells in the same row
    row_number = ws.max_row
    ws.cell(row=row_number, column=len(data) + 1).value = "X values"
    for i, x_value in enumerate(x_values):
        ws.cell(row=row_number, column=len(data) + 1 + i + 1).value = x_value
    ws.cell(row=row_number + 1, column=len(data) + 1).value = "Y values"
    for i, y_value in enumerate(y_values):
        ws.cell(row=row_number + 1, column=len(data) + 1 + i + 1).value = y_value

    # Save the workbook
    wb.save(filename)
    # Close the workbook
    wb.close()
